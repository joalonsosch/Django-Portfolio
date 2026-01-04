"""Management command to load portfolio data from Excel file.

This command reads datos.xlsx and loads:
- Assets (17 assets)
- Portfolios (Portfolio 1 and Portfolio 2)
- Initial weights for each asset in each portfolio
- Historical prices for all assets
"""
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from portfolios.models import Asset, Portfolio, Price, PortfolioWeight
from portfolios.services import (
    asset_create,
    portfolio_create,
    price_create,
    portfolio_weight_create,
)


def parse_date(date_str: str) -> date:
    """Parse date from DD/MM/YY format.
    
    Args:
        date_str: Date string in format "DD/MM/YY" (e.g., "15/02/22")
    
    Returns:
        date object
    
    Raises:
        ValueError: If date format is invalid
    """
    try:
        parts = date_str.split('/')
        if len(parts) != 3:
            raise ValueError(f"Invalid date format: {date_str}")
        
        day = int(parts[0])
        month = int(parts[1])
        year = int(parts[2])
        
        # Handle 2-digit years: assume 20XX for years 00-99
        if year < 100:
            year += 2000
        
        return date(year, month, day)
    except (ValueError, IndexError) as e:
        raise ValueError(f"Invalid date format: {date_str}") from e


class Command(BaseCommand):
    help = 'Load portfolio data from datos.xlsx Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='data/datos.xlsx',
            help='Path to Excel file (default: data/datos.xlsx)'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear existing data before loading'
        )

    def handle(self, *args, **options):
        file_path = Path(options['file'])
        
        # Resolve relative to project root if needed
        if not file_path.is_absolute():
            project_root = Path(__file__).resolve().parent.parent.parent.parent
            file_path = project_root / file_path
        
        if not file_path.exists():
            raise CommandError(f'Excel file not found: {file_path}')
        
        if options['clear']:
            self.stdout.write(self.style.WARNING('Clearing existing data...'))
            Price.objects.all().delete()
            PortfolioWeight.objects.all().delete()
            Portfolio.objects.all().delete()
            Asset.objects.all().delete()
            self.stdout.write(self.style.SUCCESS('Existing data cleared.'))
        
        try:
            self.stdout.write(f'Loading Excel file: {file_path}')
            workbook = load_workbook(file_path, data_only=True)
            
            # Verify required sheets exist
            required_sheets = ['weights', 'Precios']
            for sheet_name in required_sheets:
                if sheet_name not in workbook.sheetnames:
                    raise CommandError(f'Required sheet "{sheet_name}" not found in Excel file')
            
            weights_sheet = workbook['weights']
            precios_sheet = workbook['Precios']
            
            # Load assets
            assets = self._load_assets(weights_sheet)
            
            # Load portfolios
            portfolios = self._load_portfolios()
            
            # Load weights
            self._load_weights(weights_sheet, assets, portfolios)
            
            # Load prices
            self._load_prices(precios_sheet, assets)
            
            self.stdout.write(self.style.SUCCESS('\nData loading completed successfully!'))
            self._print_summary()
            
        except Exception as e:
            raise CommandError(f'Error loading data: {str(e)}') from e

    def _load_assets(self, weights_sheet) -> dict[str, Asset]:
        """Load assets from Weights sheet.
        
        Args:
            weights_sheet: openpyxl worksheet for Weights
        
        Returns:
            Dictionary mapping asset names to Asset instances
        """
        self.stdout.write('\nLoading assets...')
        assets = {}
        
        # Read asset names from Column B (index 2) - rows 2 onwards, skipping header row 1
        # Structure: Row 1 = ['Fecha', 'activos', 'portafolio 1', 'portafolio 2']
        #           Row 2+ = [date, asset_name, weight1, weight2]
        max_row = weights_sheet.max_row
        for row_num in range(2, max_row + 1):
            cell = weights_sheet.cell(row=row_num, column=2)  # Column B
            asset_name = cell.value
            
            if asset_name is None or not str(asset_name).strip():
                continue
            
            asset_name = str(asset_name).strip()
            
            try:
                asset = asset_create(name=asset_name)
                assets[asset_name] = asset
                self.stdout.write(f'  Created asset: {asset_name}')
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f'  Error creating asset {asset_name}: {str(e)}')
                )
                raise
        
        if len(assets) != 17:
            self.stdout.write(
                self.style.WARNING(
                    f'Warning: Expected 17 assets, found {len(assets)}'
                )
            )
        
        self.stdout.write(self.style.SUCCESS(f'Loaded {len(assets)} assets'))
        return assets

    def _load_portfolios(self) -> dict[str, Portfolio]:
        """Load portfolios (Portfolio 1 and Portfolio 2).
        
        Returns:
            Dictionary mapping portfolio names to Portfolio instances
        """
        self.stdout.write('\nLoading portfolios...')
        portfolios = {}
        
        initial_value = Decimal('1000000000.00')
        initial_date = date(2022, 2, 15)
        
        for portfolio_name in ['Portfolio 1', 'Portfolio 2']:
            try:
                portfolio = portfolio_create(
                    name=portfolio_name,
                    initial_value=initial_value,
                    initial_date=initial_date
                )
                portfolios[portfolio_name] = portfolio
                self.stdout.write(f'  Created portfolio: {portfolio_name}')
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f'  Error creating portfolio {portfolio_name}: {str(e)}')
                )
                raise
        
        self.stdout.write(self.style.SUCCESS(f'Loaded {len(portfolios)} portfolios'))
        return portfolios

    def _load_weights(
        self,
        weights_sheet,
        assets: dict[str, Asset],
        portfolios: dict[str, Portfolio]
    ):
        """Load weights from Weights sheet.
        
        Args:
            weights_sheet: openpyxl worksheet for Weights
            assets: Dictionary mapping asset names to Asset instances
            portfolios: Dictionary mapping portfolio names to Portfolio instances
        """
        self.stdout.write('\nLoading weights...')
        weights_loaded = 0
        
        # Column C = Portfolio 1 weights, Column D = Portfolio 2 weights
        # Weights are already in decimal format (0.28 = 28%), not percentages
        portfolio_columns = {
            'Portfolio 1': 3,  # Column C
            'Portfolio 2': 4,  # Column D
        }
        
        max_row = weights_sheet.max_row
        for row_num in range(2, max_row + 1):
            asset_name_cell = weights_sheet.cell(row=row_num, column=2)  # Column B
            asset_name = asset_name_cell.value
            
            if asset_name is None or not str(asset_name).strip():
                continue
            
            asset_name = str(asset_name).strip()
            
            if asset_name not in assets:
                self.stdout.write(
                    self.style.WARNING(f'  Asset {asset_name} not found, skipping')
                )
                continue
            
            asset = assets[asset_name]
            
            for portfolio_name, col_num in portfolio_columns.items():
                if portfolio_name not in portfolios:
                    continue
                
                portfolio = portfolios[portfolio_name]
                weight_cell = weights_sheet.cell(row=row_num, column=col_num)
                weight_value = weight_cell.value
                
                if weight_value is None:
                    continue
                
                try:
                    # Weights are already in decimal format (0.28 = 28%)
                    if isinstance(weight_value, (int, float)):
                        weight_decimal = Decimal(str(weight_value))
                    else:
                        weight_decimal = Decimal(str(weight_value))
                    
                    # Validate weight range (0 to 1)
                    if weight_decimal < 0 or weight_decimal > 1:
                        self.stdout.write(
                            self.style.WARNING(
                                f'  Invalid weight {weight_decimal} for {asset_name} in {portfolio_name}, skipping'
                            )
                        )
                        continue
                    
                    portfolio_weight_create(
                        portfolio=portfolio,
                        asset=asset,
                        initial_weight=weight_decimal
                    )
                    weights_loaded += 1
                    
                except (ValueError, InvalidOperation) as e:
                    self.stdout.write(
                        self.style.WARNING(
                            f'  Invalid weight value for {asset_name} in {portfolio_name}: {str(e)}'
                        )
                    )
        
        self.stdout.write(self.style.SUCCESS(f'Loaded {weights_loaded} weights'))

    def _load_prices(self, precios_sheet, assets: dict[str, Asset]):
        """Load prices from Precios sheet.
        
        Args:
            precios_sheet: openpyxl worksheet for Precios
            assets: Dictionary mapping asset names to Asset instances
        """
        self.stdout.write('\nLoading prices...')
        prices_loaded = 0
        
        # Row 1 contains asset names as column headers
        header_row = 1
        asset_columns = {}  # Map asset name to column number
        
        # Find asset columns (starting from column B, which is column 2)
        max_col = precios_sheet.max_column
        for col_num in range(2, max_col + 1):
            cell = precios_sheet.cell(row=header_row, column=col_num)
            asset_name = cell.value
            
            if asset_name is None or not str(asset_name).strip():
                continue
            
            asset_name = str(asset_name).strip()
            
            if asset_name in assets:
                asset_columns[col_num] = assets[asset_name]
            else:
                self.stdout.write(
                    self.style.WARNING(f'  Asset {asset_name} in header not found in assets, skipping column')
                )
        
        # Read price data (starting from row 2, column A contains dates)
        max_row = precios_sheet.max_row
        for row_num in range(2, max_row + 1):
            # Get date from column A
            date_cell = precios_sheet.cell(row=row_num, column=1)
            date_value = date_cell.value
            
            if date_value is None:
                continue
            
            # Parse date (dates are datetime objects from Excel)
            try:
                from datetime import datetime
                # Convert to date object explicitly
                if isinstance(date_value, date) and not isinstance(date_value, datetime):
                    price_date = date_value
                elif isinstance(date_value, datetime):
                    price_date = date_value.date()
                elif hasattr(date_value, 'date') and callable(getattr(date_value, 'date', None)):
                    price_date = date_value.date()
                elif isinstance(date_value, str):
                    price_date = parse_date(date_value)
                else:
                    # Try to convert to string and parse
                    price_date = parse_date(str(date_value))
                
                # Final check: ensure we have a date object (not datetime)
                if isinstance(price_date, datetime):
                    price_date = price_date.date()
                elif not isinstance(price_date, date):
                    raise ValueError(f"Could not convert to date: {date_value} (type: {type(date_value)})")
                    
            except (ValueError, AttributeError, TypeError) as e:
                self.stdout.write(
                    self.style.WARNING(f'  Invalid date in row {row_num}: {str(e)}')
                )
                continue
            
            # Validate date range (ensure both are date objects)
            min_date = date(2022, 2, 15)
            max_date = date(2023, 2, 16)
            # Double-check price_date is a date object
            if not isinstance(price_date, date):
                self.stdout.write(
                    self.style.WARNING(f'  Date type error in row {row_num}: {type(price_date)}, skipping')
                )
                continue
            if price_date < min_date or price_date > max_date:
                self.stdout.write(
                    self.style.WARNING(f'  Date {price_date} out of expected range, skipping')
                )
                continue
            
            # Read prices for each asset
            for col_num, asset in asset_columns.items():
                price_cell = precios_sheet.cell(row=row_num, column=col_num)
                price_value = price_cell.value
                
                if price_value is None:
                    continue
                
                try:
                    if isinstance(price_value, (int, float)):
                        price_decimal = Decimal(str(price_value))
                    else:
                        price_decimal = Decimal(str(price_value))
                    
                    # Validate price is positive
                    if price_decimal <= 0:
                        self.stdout.write(
                            self.style.WARNING(
                                f'  Invalid price {price_decimal} for {asset.name} on {price_date}, skipping'
                            )
                        )
                        continue
                    
                    price_create(
                        asset=asset,
                        date=price_date,
                        price=price_decimal
                    )
                    prices_loaded += 1
                    
                except (ValueError, InvalidOperation) as e:
                    self.stdout.write(
                        self.style.WARNING(
                            f'  Invalid price value for {asset.name} on {price_date}: {str(e)}'
                        )
                    )
        
        self.stdout.write(self.style.SUCCESS(f'Loaded {prices_loaded} prices'))

    def _print_summary(self):
        """Print summary of loaded data."""
        self.stdout.write('\n' + '=' * 50)
        self.stdout.write('Data Summary:')
        self.stdout.write('=' * 50)
        self.stdout.write(f'Assets: {Asset.objects.count()}')
        self.stdout.write(f'Portfolios: {Portfolio.objects.count()}')
        self.stdout.write(f'Weights: {PortfolioWeight.objects.count()}')
        self.stdout.write(f'Prices: {Price.objects.count()}')
        self.stdout.write('=' * 50)

